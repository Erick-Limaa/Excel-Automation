from openpyxl import load_workbook

archive = load_workbook('Alunos.xlsx')

# Ver as abas
print(archive.sheetnames)

# Pegar a aba ativa
act_sheet =  archive.active
print(act_sheet)

# Selecionar uam aba específica
sheet= archive['Planilha1']
print(sheet)

# Selecionar células
print(sheet['A1'].value) 
print(sheet.cell(row=1, column=2).value)

# Editar células
sheet.cell(row=1, column=2).value= 'Prova 1'
archive.save('Alunos.xlsx')

# Última linha
print(sheet.max_row)
# Ou
print(len(sheet['A']))

# Última coluna
print(sheet.max_column)
# Ou
print(len(sheet['1']))