# Percorrer toda a nossa base de dados
# Para cada item
    # Ver se o bairro já existe em uma aba, se não existir, criar aquela aba
    # Copiar os valores daquela linha e colocar na aba do bairro correspondente

from openpyxl import load_workbook
from copy import copy

def add_sheet(bairro, archive_bairros, header_style):
    if bairro not in archive_bairros.sheetnames:
        archive_bairros.create_sheet(bairro)
        new_sheet = archive_bairros[bairro]
        new_sheet['A1'].value = 'Data de Nascimento'
        new_sheet['B1'].value = 'Pessoa'
        new_sheet['C1'].value = 'Bairro'
        new_sheet['A1']._style = header_style
        new_sheet['B1']._style = header_style
        new_sheet['C1']._style = header_style

def transfer_data(OG_sheet, dest_sheet, OG_linha):
    dest_row = dest_sheet.max_row +1 
    for col in range(1, 4):
        OG_cell = OG_sheet.cell(row = OG_linha, column = col)
        dest_cell = dest_sheet.cell(row = dest_row, column = col)
        dest_cell.value = OG_cell.value
        dest_cell._style = copy(OG_cell._style)

archive_bairros = load_workbook('Bairros.xlsx')

print(archive_bairros.sheetnames)

sheet_db = archive_bairros['Base de Dados']

last_row = sheet_db.max_row
print(last_row)

style_header = copy(sheet_db['A1']._style)

for linha in range(2, last_row +1):
    bairro = sheet_db.cell(row = linha, column = 3).value
    if not bairro:
        break
    # Criar uma aba pro bairro
    add_sheet(bairro, archive_bairros, style_header)

    # Transferir as informações para a aba
    dest_sheet = archive_bairros[bairro]
    transfer_data(sheet_db, dest_sheet, linha)

archive_bairros.save('Bairros2.xlsx')